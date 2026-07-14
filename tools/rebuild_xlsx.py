# -*- coding: utf-8 -*-
"""由 data/organisations.csv 重新產生 data/Tender_Organisations.xlsx。

用法:python tools/rebuild_xlsx.py
- 有「備註」內容嘅行(即研究新增嘅機構)會以綠色 highlight。
"""
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(REPO, "data", "organisations.csv")
XLSX_PATH = os.path.join(REPO, "data", "Tender_Organisations.xlsx")

HEADERS = ["機構名稱", "招標專頁網址", "分類", "連結狀態", "最近期活動紀錄", "備註"]


def main():
    with open(CSV_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "list"
    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1A73E8")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")

    new_fill = PatternFill("solid", fgColor="E8F5E9")
    for row in rows:
        row = (row + [""] * len(HEADERS))[: len(HEADERS)]
        ws.append(row)
        if row[5].strip():
            for c in ws[ws.max_row]:
                c.fill = new_fill

    widths = [36, 60, 14, 10, 46, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    wb.save(XLSX_PATH)
    print(f"已重建 {XLSX_PATH}({len(rows)} 個機構)")


if __name__ == "__main__":
    main()
